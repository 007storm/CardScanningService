from shutil import copyfile
from openpyxl import load_workbook
from flask import Flask
from flask import request, jsonify
from flask_restful import reqparse
import re
import time
import numpy as np
import uuid
from config import *
from apphelper.image import union_rbox, base64_to_PIL
from application import trainTicket, idcard

filelock = 'file.lock'
if os.path.exists(filelock):
    os.remove(filelock)

if yoloTextFlag == 'keras' or AngleModelFlag == 'tf' or ocrFlag == 'keras':
    if GPU:
        os.environ["CUDA_VISIBLE_DEVICES"] = str(GPUID)
        import tensorflow as tf
        from keras import backend as K

        config = tf.ConfigProto()
        config.gpu_options.allocator_type = 'BFC'
        config.gpu_options.per_process_gpu_memory_fraction = 0.3  ## GPU最大占用量
        config.gpu_options.allow_growth = True  ##GPU是否可动态增加
        K.set_session(tf.Session(config=config))
        K.get_session().run(tf.global_variables_initializer())

    else:
        # CPU启动
        os.environ["CUDA_VISIBLE_DEVICES"] = ''

if yoloTextFlag == 'opencv':
    scale, maxScale = IMGSIZE
    from text.opencv_dnn_detect import text_detect
elif yoloTextFlag == 'darknet':
    scale, maxScale = IMGSIZE
    from text.darknet_detect import text_detect
elif yoloTextFlag == 'keras':
    scale, maxScale = IMGSIZE[0], 2048
    from text.keras_detect import text_detect
else:
    print("err,text engine in keras\opencv\darknet")

from text.opencv_dnn_detect import angle_detect

if ocr_redis:
    # 多任务并发识别
    from apphelper.redisbase import redisDataBase

    ocr = redisDataBase().put_values
else:
    from crnn.keys import alphabetChinese, alphabetEnglish

    if ocrFlag == 'keras':
        from crnn.network_keras import CRNN

        if chineseModel:
            alphabet = alphabetChinese
            if LSTMFLAG:
                ocrModel = ocrModelKerasLstm
            else:
                ocrModel = ocrModelKerasDense
        else:
            ocrModel = ocrModelKerasEng
            alphabet = alphabetEnglish
            LSTMFLAG = True

    elif ocrFlag == 'torch':
        from crnn.network_torch import CRNN

        if chineseModel:
            alphabet = alphabetChinese
            if LSTMFLAG:
                ocrModel = ocrModelTorchLstm
            else:
                ocrModel = ocrModelTorchDense

        else:
            ocrModel = ocrModelTorchEng
            alphabet = alphabetEnglish
            LSTMFLAG = True
    elif ocrFlag == 'opencv':
        from crnn.network_dnn import CRNN

        ocrModel = ocrModelOpencv
        alphabet = alphabetChinese
    else:
        print("err,ocr engine in keras\opencv\darknet")

    nclass = len(alphabet) + 1
    if ocrFlag == 'opencv':
        crnn = CRNN(alphabet=alphabet)
    else:
        crnn = CRNN(32, 1, nclass, 256, leakyRelu=False, lstmFlag=LSTMFLAG, GPU=GPU, alphabet=alphabet)
    if os.path.exists(ocrModel):
        crnn.load_weights(ocrModel)
    else:
        print("download model or tranform model with tools!")

    ocr = crnn.predict_job

from main import TextOcrModel

model = TextOcrModel(ocr, text_detect, angle_detect)

billList = ['通用OCR', '火车票', '身份证']

# 配置Excel文件夹路径
app = Flask(__name__, static_folder='/home/chineseocr/chineseocr/excel')


@app.route('/a39b1f62-afb4-11ea-9711-08d23ee9ec19/getExcel', methods=['POST'])
def getExcel():
    data = request.data
    parser = reqparse.RequestParser()
    parser.add_argument('name')
    parser.add_argument('position')
    parser.add_argument('local')
    parser.add_argument('company')
    parser.add_argument('email')
    parser.add_argument('phone')
    parser.add_argument('other')
    parser.add_argument('nameExcel')
    args = parser.parse_args()
    if not args:
        return jsonify({
            "code": "0000",
            "msg": "未传值"
        })
    # data = json.loads(data)
    name = args['name']
    position = args['position']
    local = args['local']
    company = args['company']
    email = args['email']
    phone = args['phone']
    other = args['other']
    nameExcel = args['nameExcel'] if args["nameExcel"] else '未命名'
    filepath = r'/home/chineseocr/excel'
    # filepath = r'C:\Users\11964\Desktop\chineseocr\excel'
    source = os.path.join(filepath, 'file.xlsx')
    target = os.path.join(r'/home/chineseocr/chineseocr/excel', f'{nameExcel}.xlsx')
    copyfile(source, target)
    wb = load_workbook(target)
    ws = wb.active
    ws.cell(row=1, column=2, value=name)
    ws.cell(row=2, column=2, value=position)
    ws.cell(row=3, column=2, value=company)
    ws.cell(row=4, column=2, value=local)
    ws.cell(row=5, column=2, value=email)
    ws.cell(row=6, column=2, value=phone)
    ws.cell(row=7, column=2, value=other)
    wb.save(target)
    # filename = f"{nameExcel}.xlsx"
    return jsonify({
        "code": "0000",
        "msg": "成功",
        "data": {
            "fileName": nameExcel,
            "filePath": "excel",
            "format": ".xlsx"
        }
    })


@app.route('/a39b1f62-afb4-11ea-9711-08d23ee9ec19/upload', methods=['POST'])
def upload():
    t = time.time()
    parser = reqparse.RequestParser()
    parser.add_argument('file')
    args = parser.parse_args()
    # data = request.json
    uidJob = uuid.uuid1().__str__()
    file = args['file']
    # data = json.loads(data)
    billModel = '通用OCR'
    textAngle = True  # 文字检测
    imgString = file.encode().split(b';base64,')[-1]
    img = base64_to_PIL(imgString)
    if img is not None:
        img = np.array(img)
    else:
        return 0

    H, W = img.shape[:2]

    while time.time() - t <= TIMEOUT:
        if os.path.exists(filelock):
            continue
        else:
            with open(filelock, 'w') as f:
                f.write(uidJob)
            detectAngle = textAngle
            result, angle = model.model(img,
                                        scale=scale,
                                        maxScale=maxScale,
                                        detectAngle=detectAngle,  # 是否进行文字方向检测，通过web传参控制
                                        MAX_HORIZONTAL_GAP=100,  # 字符之间的最大间隔，用于文本行的合并
                                        MIN_V_OVERLAPS=0.6,
                                        MIN_SIZE_SIM=0.6,
                                        TEXT_PROPOSALS_MIN_SCORE=0.1,
                                        TEXT_PROPOSALS_NMS_THRESH=0.3,
                                        TEXT_LINE_NMS_THRESH=0.99,  # 文本行之间测iou值
                                        LINE_MIN_SCORE=0.1,
                                        leftAdjustAlph=0.01,  # 对检测的文本行进行向左延伸
                                        rightAdjustAlph=0.01,  # 对检测的文本行进行向右延伸
                                        )

            if billModel == '' or billModel == '通用OCR':
                result = union_rbox(result, 0.2)
                res = {'name': '', 'position': '', 'company': '', 'local': '', 'email': '', 'phone': '',
                       'other': ''}
                # 对公司进行提取
                companyKeyword = ['公司', 'company', '银行', 'Bank', '集团', 'Group', '商行', 'Factory', 'CO', 'LTD']
                for i, x in enumerate(result):
                    for ck in companyKeyword:
                        if ck in x['text']:
                            res['company'] = x['text']
                            result.pop(i)
                            break
                    if res['company']:
                        break
                name = ''
                name_height = 0
                pop_data = 0
                # 对姓名进行处理
                for i, x in enumerate(result):
                    if 'name' in x['text'] or '姓名' in x['text']:
                        res['name'] = x['text'].replace('name', '')
                        res['name'] = res['name'].replace('姓名', '')
                        res['name'] = res['name'].replace(':', '')
                        result.pop(i)
                        break
                    else:
                        # 正则取中文，进行分别进行中文名和英文名的判断
                        text_list = re.findall(r'[\u4e00-\u9fa5]', x['text'])
                        if text_list:
                            if name_height < x['h'] and 1 < len(text_list) < 5:
                                name_height = x['h']
                                name = x['text']
                                pop_data = i
                        else:
                            strUpper = x['text'].upper()
                            if name_height < x['h'] and strUpper != x['text']:
                                name_height = x['h']
                                name = x['text']
                                pop_data = i
                    res['name'] = name
                result.pop(pop_data)
                # 对职位进行提取
                pos = ['GM', 'VP', 'HRD', 'OD', 'MD', 'OM', 'PM', 'BM', 'DM', 'RM', 'AAD', 'ACD', 'AD', 'AE',
                       'AP', 'ASM', 'VC', 'CWO', 'COO', 'CXO', 'CYO', 'CZO', 'PS', 'Manager', 'Engineer']
                posNum = "".join(filter(str.isdigit, x['text']))
                for i, x in enumerate(result[pop_data:pop_data + 3]):
                    if '职位' in x['text'] or 'position' in x['text']:
                        res['position'] = x['text'].replace('职位', '')
                        res['position'] = res['position'].replace('position', '')
                        res['position'] = res['position'].replace(':', '')
                        result.pop(i)
                        break
                if not res['position']:
                    for i, x in enumerate(result[pop_data:pop_data + 3]):
                        for ch in x['text']:
                            if u'\u4e00' <= ch <= u'\u9fff' and len(posNum) == 0:
                                res['position'] = x['text']
                                result.pop(i)
                                break
                        if res['position']:
                            break
                        for p in pos:
                            if p in x['text']:
                                res['position'] = x['text']
                                result.pop(i)
                                break
                        if res['position']:
                            break
                # 对邮箱进行提取
                for i, x in enumerate(result):
                    if '@' in x['text'] or '邮箱' in x['text'] or 'email' in x['text']:
                        res['email'] = x['text'].replace('邮箱', '')
                        res['email'] = res['email'].replace('email', '')
                        res['email'] = res['email'].replace(':', '')
                        result.pop(i)
                        break
                # 对地址进行提取
                localKeyword = ['市', '省', '区', '号', '路', '岛', '地址', 'sheng', 'shi', 'qu', 'hao', 'lu']
                for i, x in enumerate(result):
                    for k in localKeyword:
                        if k in x['text'] or 'Add' in x['text'] or 'add' in x['text']:
                            res['local'] = x['text'].replace('地址', '')
                            res['local'] = res['local'].replace('Add', '')
                            res['local'] = res['local'].replace(':', '')
                            result.pop(i)
                            break
                    if res['local']:
                        break

                # 对手机号码进行提取
                for i, x in enumerate(result):
                    if '-' not in x['text'] and '-' not in x['text']:
                        telephone = "".join(filter(str.isdigit, x['text']))
                        if 11 <= len(telephone) <= 13:
                            res['phone'] += x['text'].replace('电话', '') + ';'
                            res['phone'] = res['phone'].replace('Tel', '')
                            res['phone'] = res['phone'].replace('手机', '')
                            res['phone'] = res['phone'].replace(':', '')
                            result.pop(i)
                # 其他文本提取
                for i, x in enumerate(result):
                    res['other'] += x['text'] + ';'

            elif billModel == '火车票':
                res = trainTicket.trainTicket(result)
                res = res.res
                res = [{'text': res[key], 'name': key, 'box': {}} for key in res]

            elif billModel == '身份证':

                res = idcard.idcard(result)
                res = res.res
                res = [{'text': res[key], 'name': key, 'box': {}} for key in res]

            os.remove(filelock)
            break

    timeTake = time.time() - t
    return jsonify({
        "code": "0000",
        "msg": "成功",
        "data": res
    })
    # return json.dumps({'res': res, 'timeTake': round(timeTake, 4)}, ensure_ascii=False)


if __name__ == '__main__':
    app.debug = True  # 设置调试模式，生产模式的时候要关掉debug
    # CORS(app, supports_credentials=True)
    app.run(debug=True, host='0.0.0.0', port=18080)
